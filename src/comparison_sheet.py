"""
Combines multiple per-supplier extractions (one image/PDF per supplier,
each already run through src/nodes/extract_quote.py) into one comparison
view, and renders that view as a downloadable .xlsx matching the original
multi-supplier comparison sheet layout. No database involved — this is a
pure in-memory build for a download button.
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")
_THIN = Side(style="thin", color="999999")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

SUPPLIER_SUBCOLS = ["Packing", "Brand", "Origin", "Price"]


def merge_extractions(parsed_list: list) -> dict:
    """
    Combines multiple per-supplier extractions (each the {"items": [...]}
    shape returned by extract_quote()) into one {"items": [...]}
    comparison structure.

    Items are matched across files by exact, whitespace-normalized,
    case-insensitive item_description — no fuzzy matching. If two
    suppliers describe "the same" item with different wording, they show
    up as two separate rows rather than being silently merged.
    """
    items_by_key = {}
    item_order = []

    for parsed in parsed_list:
        for item in parsed.get("items", []):
            key = " ".join(item["item_description"].split()).lower()
            if key not in items_by_key:
                items_by_key[key] = {
                    "item_description": item["item_description"],
                    "yearly_consumption": item.get("yearly_consumption"),
                    "suppliers": [],
                }
                item_order.append(key)
            items_by_key[key]["suppliers"].extend(item.get("suppliers", []))

    return {"items": [items_by_key[k] for k in item_order]}


def _collect_supplier_names(items: list) -> list:
    names = []
    seen = set()
    for item in items:
        for s in item.get("suppliers", []):
            name = (s.get("supplier_name") or "").strip()
            if name and name.lower() not in seen:
                seen.add(name.lower())
                names.append(name)
    return names


def build_comparison_excel(items: list) -> bytes:
    """
    Builds an .xlsx matching the original comparison sheet layout: one
    4-column block (Packing / Brand / Origin / Price) per supplier across
    the item rows. Returns raw .xlsx bytes for a download button — nothing
    is written anywhere.
    """
    supplier_names = _collect_supplier_names(items)
    n_suppliers = len(supplier_names)
    total_cols = 1 + n_suppliers * 4

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(total_cols, 1))
    title_cell = ws.cell(row=1, column=1, value="SUPPLIER PRICE COMPARISON")
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = CENTER

    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
    c = ws.cell(row=2, column=1, value="Item Description")
    c.font, c.alignment, c.fill, c.border = BOLD, CENTER, HEADER_FILL, BORDER

    col = 2
    for name in supplier_names:
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 3)
        head = ws.cell(row=2, column=col, value=name)
        head.font, head.alignment, head.fill, head.border = BOLD, CENTER, HEADER_FILL, BORDER
        for offset, sub in enumerate(SUPPLIER_SUBCOLS):
            sub_cell = ws.cell(row=3, column=col + offset, value=sub)
            sub_cell.font, sub_cell.alignment = BOLD, CENTER
            sub_cell.fill, sub_cell.border = HEADER_FILL, BORDER
        col += 4

    row = 4
    for item in items:
        desc_cell = ws.cell(row=row, column=1, value=item["item_description"])
        desc_cell.alignment, desc_cell.border = LEFT, BORDER

        suppliers_by_name = {
            (s.get("supplier_name") or "").strip().lower(): s
            for s in item.get("suppliers", [])
        }
        col = 2
        for name in supplier_names:
            s = suppliers_by_name.get(name.lower())
            values = (
                [s.get("packing"), s.get("brand"), s.get("origin"), s.get("unit_price")]
                if s else ["", "", "", ""]
            )
            for offset, val in enumerate(values):
                cell = ws.cell(row=row, column=col + offset, value=val)
                cell.border = BORDER
                cell.alignment = CENTER if offset == 3 else LEFT
            col += 4
        row += 1

    ws.column_dimensions["A"].width = 32
    for i in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
